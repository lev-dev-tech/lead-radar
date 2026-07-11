import asyncio
import hashlib
import os
import re
import sys
from datetime import datetime, timedelta, timezone

from telethon import TelegramClient, functions
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

import config

# Выбор аккаунта: локально — основной (папка «Вакансии»), в облаке (GitHub Actions)
# — расходный (RADAR_ACCOUNT=burner), он читает каналы из config.CHANNELS.
if os.getenv("RADAR_ACCOUNT", "lev").lower() == "burner":
    SESSION, ACC_API_ID, ACC_API_HASH = (
        config.BURNER_SESSION, config.BURNER_API_ID, config.BURNER_API_HASH)
else:
    SESSION, ACC_API_ID, ACC_API_HASH = (
        config.LEV_SESSION, config.LEV_API_ID, config.LEV_API_HASH)

RE_USERNAME = re.compile(r"@([A-Za-z][A-Za-z0-9_]{3,31})")
RE_PHONE = re.compile(r"(?:\+?7|8)[\s\-(]*\d{3}[\s\-)]*\d{3}[\s\-]*\d{2}[\s\-]*\d{2}")
RE_TME = re.compile(r"(?:https?://)?t\.me/([A-Za-z][A-Za-z0-9_]{3,31})")
RE_HASHTAG = re.compile(r"#([0-9a-zA-Zа-яёА-ЯЁ_]+)")
RE_NONALPHA = re.compile(r"[^а-яёa-z0-9]")

EXCLUDE_TAGS = {t.lower() for t in config.EXCLUDE_HASHTAGS}
VAC_TAGS = {t.lower() for t in config.VACANCY_HASHTAGS}


def low(s):
    return (s or "").lower()


def contains_any(text, keywords):
    t = low(text)
    return [k for k in keywords if k in t]


def get_hashtags(text):
    return {"#" + h.lower() for h in RE_HASHTAG.findall(text or "")}


def text_fingerprint(text):
    normalized = RE_NONALPHA.sub("", low(text))[:200]
    return hashlib.md5(normalized.encode()).hexdigest()


def is_vacancy(text):
    if not text:
        return False, 0, ""

    tags = get_hashtags(text)

    if tags & EXCLUDE_TAGS:
        return False, 0, f"хэштег исполнителя"

    topic_stop = contains_any(text, config.TOPIC_EXCLUDE)
    if topic_stop:
        return False, 0, f"не та тема: {topic_stop}"

    strong = contains_any(text, config.STRONG_PROMO_KEYWORDS)
    if strong:
        return False, 0, f"самореклама"

    # Тех-тема обязательна всегда — иначе общий хэштег #ищу/#вакансия
    # пропускал бы весь инфобиз-мусор ("куратор", "удалёнка для девушек").
    tech = contains_any(text, config.TECH_KEYWORDS)
    if not tech:
        return False, 0, "нет тех-темы"

    vac_tags = tags & VAC_TAGS
    vac = contains_any(text, config.VACANCY_KEYWORDS)
    if not (vac_tags or vac):
        return False, 0, "нет признаков вакансии"

    soft = contains_any(text, config.SELF_PROMO_KEYWORDS)
    if len(soft) > len(vac) + len(vac_tags):
        return False, 0, f"больше саморекламы"

    score = len(tech) + len(vac) + (10 if vac_tags else 0)
    return True, score, f"tech={tech[:2]}; vac={vac[:2]}"


def extract_contact(text, sender):
    parts = []
    for m in RE_USERNAME.findall(text or ""):
        parts.append("@" + m)
    for m in RE_TME.findall(text or ""):
        u = "@" + m
        if u not in parts:
            parts.append(u)
    parts.extend(RE_PHONE.findall(text or ""))

    sender_str = ""
    if sender is not None:
        name = " ".join(filter(None, [getattr(sender, "first_name", None),
                                       getattr(sender, "last_name", None)]))
        uname = getattr(sender, "username", None)
        bits = []
        if name:
            bits.append(name.strip())
        if uname:
            bits.append("@" + uname)
        sender_str = " / ".join(bits)

    text_contacts = ", ".join(dict.fromkeys(parts))
    if text_contacts and sender_str:
        return f"{text_contacts}  (отправитель: {sender_str})"
    return text_contacts or sender_str or "(не указан)"


def peer_id(peer):
    for attr in ("channel_id", "chat_id", "user_id"):
        v = getattr(peer, attr, None)
        if v is not None:
            return v
    return None


def chat_name_of(entity, peer):
    if entity is not None:
        return (getattr(entity, "title", None) or getattr(entity, "username", None)
                or " ".join(filter(None, [getattr(entity, "first_name", None),
                                          getattr(entity, "last_name", None)]))
                or f"chat {entity.id}")
    return f"chat {peer_id(peer)}"


def build_link(entity, peer, msg_id):
    uname = getattr(entity, "username", None)
    if uname:
        return f"https://t.me/{uname}/{msg_id}"
    cid = getattr(entity, "id", None) or peer_id(peer)
    if cid is not None:
        return f"https://t.me/c/{cid}/{msg_id}"
    return ""


def title_text(t):
    if t is None:
        return ""
    if isinstance(t, str):
        return t
    return getattr(t, "text", "") or str(t)


async def get_folder_peers(client):
    res = await client(functions.messages.GetDialogFiltersRequest())
    filters = getattr(res, "filters", res)
    target = None
    found_names = []
    for f in filters:
        name = title_text(getattr(f, "title", None))
        if name:
            found_names.append(name)
        if name and name.strip().lower() == config.FOLDER_NAME.strip().lower():
            target = f
    if target is None:
        print(f"[!] Папка '{config.FOLDER_NAME}' не найдена.")
        print("    Доступные папки:", ", ".join(found_names) or "(нет)")
        return None
    peers = getattr(target, "include_peers", []) or []
    print(f"[+] Папка '{config.FOLDER_NAME}': {len(peers)} чатов")
    return peers


async def main():
    if not ACC_API_ID or not ACC_API_HASH:
        print("[ОШИБКА] Не заданы API_ID/API_HASH для выбранного аккаунта")
        sys.exit(1)

    client = TelegramClient(SESSION, ACC_API_ID, ACC_API_HASH)
    await client.connect()
    if not await client.is_user_authorized():
        print(f"[ОШИБКА] Сессия {SESSION} не авторизована.")
        await client.disconnect()
        sys.exit(1)
    me = await client.get_me()
    print(f"[+] Авторизация успешна: @{me.username}")

    peers = await get_folder_peers(client)
    if not peers:
        channels = getattr(config, "CHANNELS", [])
        if channels:
            print(f"[+] Папки нет — беру {len(channels)} каналов из config.CHANNELS")
            peers = channels
        else:
            await client.disconnect()
            return

    cutoff = datetime.now(timezone.utc) - timedelta(days=config.DAYS_BACK)
    results = []
    seen_fingerprints = set()
    scanned = 0
    failed = 0
    total_msgs = 0
    dupes = 0

    for peer in peers:
        entity = None
        try:
            entity = await client.get_entity(peer)
        except Exception:
            entity = None

        name = chat_name_of(entity, peer)
        target = entity if entity is not None else peer

        try:
            chat_msgs = 0
            async for msg in client.iter_messages(target, limit=config.MAX_MESSAGES_PER_CHAT):
                if msg.date < cutoff:
                    break
                chat_msgs += 1
                text = msg.message or ""
                if not text:
                    continue

                fp = text_fingerprint(text)
                if fp in seen_fingerprints:
                    dupes += 1
                    continue
                seen_fingerprints.add(fp)

                ok, score, _ = is_vacancy(text)
                if not ok:
                    continue

                try:
                    sender = await msg.get_sender()
                except Exception:
                    sender = None

                results.append({
                    "chat": name,
                    "contact": extract_contact(text, sender),
                    "link": build_link(entity, peer, msg.id),
                    "date": msg.date.astimezone(),
                    "date_str": msg.date.astimezone().strftime("%Y-%m-%d %H:%M"),
                    "text": text[:600],
                })

            total_msgs += chat_msgs
            scanned += 1
            print(f"[{scanned}/{len(peers)}] {name} — сообщений: {chat_msgs}")
        except Exception as e:
            failed += 1
            print(f"    [skip] {name}: {e}")
            continue

    await client.disconnect()
    print(f"\n[+] Чатов: {scanned}, ошибок: {failed}, "
          f"сообщений: {total_msgs}, дублей отсеяно: {dupes}")
    print(f"[+] Вакансий найдено: {len(results)}")
    save_xlsx(results)


def save_xlsx(results):
    results.sort(key=lambda r: r["date"], reverse=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Вакансии"
    headers = ["№", "Контакт заказчика", "Ссылка на сообщение", "Чат", "Дата", "Текст"]
    ws.append(headers)

    head_fill = PatternFill("solid", fgColor="2E86C1")
    head_font = Font(bold=True, color="FFFFFF")
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(vertical="center", horizontal="center")

    for i, r in enumerate(results, 1):
        ws.append([i, r["contact"], r["link"], r["chat"], r["date_str"], r["text"]])
        link_cell = ws.cell(row=i + 1, column=3)
        if r["link"]:
            link_cell.hyperlink = r["link"]
            link_cell.font = Font(color="1A5276", underline="single")

    widths = [5, 38, 40, 28, 16, 65]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(config.OUTPUT_FILE)
    print(f"[+] Таблица сохранена: {config.OUTPUT_FILE}")


if __name__ == "__main__":
    asyncio.run(main())
